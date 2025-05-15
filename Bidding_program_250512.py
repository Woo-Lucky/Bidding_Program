import sys
import requests
import pandas as pd
import os
import webbrowser
from openpyxl import load_workbook
from PyQt5.QtWidgets import QApplication, QDialog, QMessageBox
from PyQt5.QtCore import QTimer, QTime
from PyQt5 import uic
from datetime import datetime, timedelta

base_dir = os.path.dirname(os.path.abspath(__file__))
ui_path = "uitest_250512.ui" # Qt Designer로 만든 ui 파일

# 공고 종류별 API URL
open_api_Cnstwk = 'http://apis.data.go.kr/1230000/ad/BidPublicInfoService/getBidPblancListInfoCnstwkPPSSrch' # 나라장터검색조건에 의한 입찰공고공사조회
open_api_Servc = 'http://apis.data.go.kr/1230000/ad/BidPublicInfoService/getBidPblancListInfoServcPPSSrch'   # 나라장터검색조건에 의한 입찰공고용역조회
open_api_Frgcpt = 'http://apis.data.go.kr/1230000/ad/BidPublicInfoService/getBidPblancListInfoFrgcptPPSSrch' # 나라장터검색조건에 의한 입찰공고외자조회
open_api_Thng = 'http://apis.data.go.kr/1230000/ad/BidPublicInfoService/getBidPblancListInfoThngPPSSrch'     # 나라장터검색조건에 의한 입찰공고물품조회
open_api_Etc = 'http://apis.data.go.kr/1230000/ad/BidPublicInfoService/getBidPblancListInfoEtcPPSSrch'       # 나라장터검색조건에 의한 입찰공고기타조회

class MainDialog(QDialog):
    def __init__(self):
        super().__init__()
        uic.loadUi(os.path.join(base_dir, ui_path), self)
        self.keyword_edit.setPlaceholderText("예: 조경, 비탈면 등 (콤마로 구분)") # 키워드 입력 도우미
        self.start_date_edit.setDate(datetime.today()) # 날짜를 "오늘 날짜"로 세팅
        self.end_date_edit.setDate(datetime.today()) # 날짜를 "오늘 날짜"로 세팅
        self.start_date_edit.setStyleSheet("""
            QDateTimeEdit {
                background-color: white;
            }
            QCalendarWidget QAbstractItemView {
                background-color: white;
                selection-background-color: #0078d7;
                selection-color: white;
            }
            """)
        # 버튼 설정
        self.pushButton.clicked.connect(self.fetch_data)
        self.pushButton.clicked.connect(self.handle_save)
        self.go_bidding.clicked.connect(lambda: webbrowser.open('https://www.g2b.go.kr/'))
        self.api_data = None
        self.bidding_type.addItems([ # 공고 종류 항목 추가
            "공사",
            "용역",
            "물품",
            "외자",
            "기타"])
        self.region_limit.addItems([ # 입찰제한지역 항목 추가
            "전국",
            "서울특별시",
            "부산광역시",
            "대구광역시",
            "인천광역시",
            "광주광역시",
            "대전광역시",
            "울산광역시",
            "세종특별자치시",
            "경기도",
            "충청북도",
            "충청남도",
            "전라남도",
            "경상북도",
            "경상남도",
            "제주특별자치도",
            "강원특별자치도",
            "전북특별자치도"])
        
        # QTimer 설정 ----- 일정 시간마다 자동으로 입찰 공고 수집
        self.timer = QTimer(self)
        self.timer.timeout.connect(self.check_time_and_save)  # 타이머가 만료되면 check_time_and_save 실행
        self.timer.start(1000)  # 1초마다 실행 (1000ms)
        # 자동 실행 시간 설정
        self.target_times = [
            QTime(9, 0, 0),   # 오전 9시
            QTime(13, 0, 0),  # 오후 1시
            QTime(17, 0, 0)  # 오후 5시
        ]
        self.triggered_times = set()  # 이미 실행된 시간을 저장
        
    def check_time_and_save(self):
        """특정 시간에 handle_save 메서드를 자동 실행"""
        current_time = QTime.currentTime()

        for target_time in self.target_times:
        # 현재 시간이 target_time보다 크고, 아직 실행되지 않은 경우 실행
            if current_time >= target_time and target_time not in self.triggered_times:
                self.triggered_times.add(target_time)
                try:
                    self.fetch_data_auto()
                    self.handle_save_auto()
                except Exception as e:
                    print(f"[자동저장 오류] {target_time.toString()} 실패: {str(e)}")
        if current_time.hour() == 0 and current_time.minute() == 0 and current_time.second() < 5:
            self.triggered_times.clear()
        
    def handle_save(self): #공고 종류 별 저장 타입 변경 설정!!!
            """Handle save logic based on bidding type."""
            bidding_type = self.bidding_type.currentText()
            if bidding_type == "공사":
                self.save_file_Cnstwk()
            elif bidding_type == "용역":
                self.save_file_Servc()
            elif bidding_type == "물품":
                self.save_file_Thng()
            elif bidding_type == "외자":
                self.save_file_Frgcpt()
            elif bidding_type == "기타":
                self.save_file_Etc()
            else:
                QMessageBox.warning(self, "경고", "올바른 입찰 유형을 선택하세요.")
                
    def handle_save_auto(self): #공고 종류 별 저장 타입 변경 설정!!!
            """Handle save logic based on bidding type."""
            bidding_type = self.bidding_type.currentText()
            if bidding_type == "공사":
                self.save_file_Cnstwk_auto()
            elif bidding_type == "용역":
                self.save_file_Servc_auto()
            elif bidding_type == "물품":
                self.save_file_Thng_auto()
            elif bidding_type == "외자":
                self.save_file_Frgcpt_auto()
            elif bidding_type == "기타":
                self.save_file_Etc_auto()
            else:
                QMessageBox.warning(self, "경고", "올바른 입찰 유형을 선택하세요.")

    def fetch_data(self): # 데이터 받아오기
        """API 데이터를 가져오는 함수"""
        service_key = 'snOlXoqmO+Auj39eC+RSnF27Y7g8dM+g4YxudtLaq5664xOw1XdFW6JMGQE1Fi/3981bQGnKsooawG67Jn0zTg=='
        #setvice_key = self.service_key_edit.text() # 서비스키 수동 입력
        start_date = self.start_date_edit.text() + '0000'
        start_date = start_date.replace('-','').replace(':','').replace(' ','')
        end_date = self.end_date_edit.text() + '2359'
        end_date=end_date.replace('-','').replace(':','').replace(' ','')
        keyword = self.keyword_edit.toPlainText().split(',')
        
        # 입찰제한지역명을 지역코드로 변환
        if self.region_limit.currentText()=="서울특별시":
            region_limit=11
        elif self.region_limit.currentText()=="부산광역시":
            region_limit=26
        elif self.region_limit.currentText()=="대구광역시":
            region_limit=27
        elif self.region_limit.currentText()=="인천광역시":
            region_limit=28
        elif self.region_limit.currentText()=="광주광역시":
            region_limit=29
        elif self.region_limit.currentText()=="대전광역시":
            region_limit=30
        elif self.region_limit.currentText()=="울산광역시":
            region_limit=31
        elif self.region_limit.currentText()=="세종특별자치시":
            region_limit=36
        elif self.region_limit.currentText()=="경기도":
            region_limit=41
        elif self.region_limit.currentText()=="충청북도":
            region_limit=43            
        elif self.region_limit.currentText()=="충청남도":
            region_limit=44            
        elif self.region_limit.currentText()=="전라남도":
            region_limit=46
        elif self.region_limit.currentText()=="경상북도":
            region_limit=47            
        elif self.region_limit.currentText()=="경상남도":
            region_limit=48            
        elif self.region_limit.currentText()=="제주특별자치도":
            region_limit=50             
        elif self.region_limit.currentText()=="강원특별자치도":
            region_limit=51 
        elif self.region_limit.currentText()=="전북특별자치도":
            region_limit=52 
        
        # 입찰마감여부 체크
        if self.bidclose_yn.isChecked()==True:
            bidclose_yn = "Y"
        else: bidclose_yn = "N"

        # 콤보박스에서 선택된 항목에 따라 open_api URL 변경
        bidding_type = self.bidding_type.currentText()
        if bidding_type == "공사":
            open_api = open_api_Cnstwk
        elif bidding_type == "용역":
            open_api = open_api_Servc
        elif bidding_type == "물품":
            open_api = open_api_Thng
        elif bidding_type == "외자":
            open_api = open_api_Frgcpt
        elif bidding_type == "기타":
            open_api = open_api_Etc
        else:
            QMessageBox.warning(self, "경고", "올바른 입찰 유형을 선택하세요.")
            return
        
        all_dataframes = []
        
        for kw in keyword:
            keyword = kw.strip()
            params = { # 파라미터 설정
            'serviceKey': service_key,
            'inqryDiv': 1,
            'inqryBgnDt': start_date,
            'inqryEndDt': end_date,
            'pageNo': 1,
            'numOfRows': 999,
            'bidNtceNm': keyword,
            'prtcptLmtRgnCd':region_limit,
            'bidClseExpYn':bidclose_yn,
            'type': 'json'
            }

            try:
                response = requests.get(open_api, params=params)
                print(response.url)  # 요청 URL 출력 (디버깅용)
                if response.status_code == 200:
                    data = response.json()
                    if 'response' in data and 'body' in data['response'] and 'items' in data['response']['body']:
                        bid_data = data['response']['body']['items']
                        df = pd.DataFrame(bid_data)

                        # bidding_type(공고 종류)에 따라 각각 가져올 데이터를 열로 정리
                        if bidding_type == "공사":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'bdgtAmt', 'sucsfbidLwltRate', 'sucsfbidMthdNm', 'stdNtceDocUrl']]
                
                        elif bidding_type == "용역":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'presmptPrce', 'sucsfbidLwltRate', 'sucsfbidMthdNm']]
                
                        elif bidding_type == "물품":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'bdgtAmt', 'sucsfbidLwltRate', 'sucsfbidMthdNm']]
                
                        elif bidding_type == "외자":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'bdgtAmt', 'sucsfbidLwltRate', 'sucsfbidMthdNm']]
                
                        elif bidding_type == "기타":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'bdgtAmt', 'sucsfbidLwltRate', 'sucsfbidMthdNm']]
                    
                        self.api_data = self.api_data.rename(columns={
                        'bidNtceNo': '입찰공고번호',
                        'ntceKindNm': '공고종류명',
                        'bidNtceDt': '입찰공고일시',
                        'bidNtceNm': '입찰공고명',
                        'ntceInsttNm': '공고기관명',
                        'dminsttNm': '수요기관명',
                        'bidMethdNm': '입찰방식명',
                        'bidBeginDt': '입찰개시일시',
                        'bidClseDt': '입찰마감일시',
                        'bidPrtcptLmtYn': '입찰참가제한여부',
                        'bdgtAmt': '예산금액',
                        'presmptPrce': '추정가격',
                        'sucsfbidLwltRate': '낙찰하한율',
                        'sucsfbidMthdNm': '낙찰방법명',
                        'stdNtceDocUrl': '표준공고서URL',
                        'asignBdgtAmt': '배정예산금액',
                    })
                    
                        all_dataframes.append(self.api_data)
            except Exception as e:
                print(f"키워드 '{keyword}' 처리 중 오류:", e)
        
        if all_dataframes:
            merged_df = pd.concat(all_dataframes, ignore_index=True).drop_duplicates()
            self.api_data = merged_df            
            QMessageBox.information(self, "성공", "데이터를 성공적으로 가져왔습니다!")
        else:
            QMessageBox.warning(self, "경고", "API 응답에 데이터가 없습니다.")


    def fetch_data_auto(self): # 데이터 받아오기 자동버전
        """API 데이터를 가져오는 함수"""
        service_key = 'snOlXoqmO+Auj39eC+RSnF27Y7g8dM+g4YxudtLaq5664xOw1XdFW6JMGQE1Fi/3981bQGnKsooawG67Jn0zTg=='
        today = datetime.now()
        yesterday = today - timedelta(days=1)
        #setvice_key = self.service_key_edit.text() # 서비스키 수동 입력
        start_date = yesterday.strftime("%Y%m%d") + '0000'  # 어제 날짜로 설정
        start_date = start_date.replace('-','').replace(':','').replace(' ','')
        end_date = today.strftime("%Y%m%d") + '2359'  # 오늘 날짜로 설정
        end_date=end_date.replace('-','').replace(':','').replace(' ','')
        keyword = self.keyword_edit.toPlainText().split(',')

        # 콤보박스에서 선택된 항목에 따라 open_api URL 변경
        bidding_type = self.bidding_type.currentText()
        if bidding_type == "공사":
            open_api = open_api_Cnstwk
        elif bidding_type == "용역":
            open_api = open_api_Servc
        elif bidding_type == "물품":
            open_api = open_api_Thng
        elif bidding_type == "외자":
            open_api = open_api_Frgcpt
        elif bidding_type == "기타":
            open_api = open_api_Etc
        else:
            QMessageBox.warning(self, "경고", "올바른 입찰 유형을 선택하세요.")
            return
        
        all_dataframes = []
        for kw in keyword:
            keyword = kw.strip()
        # 파라미터 설정
            params = { 
            'serviceKey': service_key,
            'inqryDiv': 1,
            'inqryBgnDt': start_date,
            'inqryEndDt': end_date,
            'pageNo': 1,
            'numOfRows': 9999,
            'bidNtceNm': keyword,
            'type': 'json'
        }

            try:
                response = requests.get(open_api, params=params)
                if response.status_code == 200:
                    data = response.json()
                    if 'response' in data and 'body' in data['response'] and 'items' in data['response']['body']:
                        bid_data = data['response']['body']['items']
                        df = pd.DataFrame(bid_data)
                    
                    # bidding_type(공고 종류)에 따라 각각 가져올 데이터를 열로 정리
                        if bidding_type == "공사":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'bdgtAmt', 'sucsfbidLwltRate', 'sucsfbidMthdNm', 'stdNtceDocUrl']]
                
                        elif bidding_type == "용역":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'presmptPrce', 'sucsfbidLwltRate', 'sucsfbidMthdNm']]
                
                        elif bidding_type == "물품":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'bdgtAmt', 'sucsfbidLwltRate', 'sucsfbidMthdNm']]
                
                        elif bidding_type == "외자":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'bdgtAmt', 'sucsfbidLwltRate', 'sucsfbidMthdNm']]
                
                        elif bidding_type == "기타":
                            self.api_data = df[['bidNtceNo', 'ntceKindNm', 'bidNtceDt', 'bidNtceNm', 'ntceInsttNm', 'dminsttNm', 'bidMethdNm', 'bidBeginDt',
                                        'bidClseDt', 'bidPrtcptLmtYn', 'bdgtAmt', 'sucsfbidLwltRate', 'sucsfbidMthdNm']]
                    
                        self.api_data = self.api_data.rename(columns={
                        'bidNtceNo': '입찰공고번호',
                        'ntceKindNm': '공고종류명',
                        'bidNtceDt': '입찰공고일시',
                        'bidNtceNm': '입찰공고명',
                        'ntceInsttNm': '공고기관명',
                        'dminsttNm': '수요기관명',
                        'bidMethdNm': '입찰방식명',
                        'bidBeginDt': '입찰개시일시',
                        'bidClseDt': '입찰마감일시',
                        'bidPrtcptLmtYn': '입찰참가제한여부',
                        'bdgtAmt': '예산금액',
                        'presmptPrce': '추정가격',
                        'sucsfbidLwltRate': '낙찰하한율',
                        'sucsfbidMthdNm': '낙찰방법명',
                        'stdNtceDocUrl': '표준공고서URL',
                        'asignBdgtAmt': '배정예산금액',
                        })
                        
                        all_dataframes.append(self.api_data)
            except Exception as e:
                print(f"키워드 '{keyword}' 처리 중 오류:", e)

        if all_dataframes:
            merged_df = pd.concat(all_dataframes, ignore_index=True).drop_duplicates()
            self.api_data = merged_df
            # QMessageBox.information(self, "성공", "데이터를 성공적으로 가져왔습니다!")
        else:
            QMessageBox.warning(self, "경고", "API 응답에 데이터가 없습니다.")

    def save_file_Cnstwk(self): # 공사 파일 저장 변수 설정!!!
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return
        
        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_공사_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가

        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        
        #파일 경로 설정정
        file_path = os.path.join(save_dir, file_name)

        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정
                
                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")

    def save_file_Cnstwk_auto(self): # 공사 파일 저장 변수 설정!!! 자동버전
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return
        
        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_공사_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가
        file_path = os.path.join(save_dir, file_name)

        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정
                
                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                #QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")
        
    def save_file_Servc(self): # 용역 파일 저장 변수 설정!!!
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return

        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_용역_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가
        file_path = os.path.join(save_dir, file_name)
        
        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정
                
                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")
                
    def save_file_Servc_auto(self): # 용역 파일 저장 변수 설정!!! 자동버전
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return
        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_용역_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가
        file_path = os.path.join(save_dir, file_name)
        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정

                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                #QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")

    def save_file_Thng(self): # 물품 파일 저장 변수 설정!!!
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return
        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_물품_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가
        file_path = os.path.join(save_dir, file_name)
        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정
                
                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")
                
    def save_file_Thng_auto(self): # 물품 파일 저장 변수 설정!!! 자동버전
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return
        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_물품_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가
        file_path = os.path.join(save_dir, file_name)
        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정
                
                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                #QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")

    def save_file_Frgcpt(self): # 외자 파일 저장 변수 설정!!!
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return
        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_외자_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가
        file_path = os.path.join(save_dir, file_name)
        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정
                
                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")

    def save_file_Frgcpt_auto(self): # 외자 파일 저장 변수 설정!!! 자동버전
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return
        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_외자_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가
        file_path = os.path.join(save_dir, file_name)
        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정
                
                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                #QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")
  
    def save_file_Etc(self): # 기타 파일 저장 변수 설정!!!
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return
        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_기타_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가
        file_path = os.path.join(save_dir, file_name)
        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정
                
                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")

    def save_file_Etc_auto(self): # 기타 파일 저장 변수 설정!!! 자동버전
        """가져온 데이터를 엑셀 파일로 저장"""
        if self.api_data is None:
            QMessageBox.warning(self, "경고", "저장할 데이터가 없습니다. 먼저 데이터를 가져오세요.")
            return
        current_time = datetime.now().strftime("%Y%m%d_%H%M")  # 예: 20250407_1530
        # C 드라이브에 '입찰공고모음' 폴더 생성
        save_dir = r"C:\입찰공고모음"
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)  # 폴더가 없으면 생성
        file_name = f"{current_time}_{self.keyword_edit.toPlainText()}_기타_입찰공고.xlsx"  # 파일 이름에 날짜와 시간 추가
        file_path = os.path.join(save_dir, file_name)
        try:
                # 데이터를 엑셀 파일로 저장
                self.api_data.to_excel(file_path, index=False)

                # 엑셀 파일 열 크기 조정
                wb = load_workbook(file_path)
                ws = wb.active
                
                # 첫 행 고정
                ws.freeze_panes = "A2"  # A2 셀을 기준으로 첫 행 고정
                
                # 모든 열의 너비를 텍스트 길이에 맞게 조정
                for col in ws.columns:
                    max_length = 0
                    column = col[0].column_letter  # 열 이름(A, B, C 등)
                    for cell in col:
                        try:
                            if cell.value:  # 셀 값이 있는 경우
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass
                    adjusted_width = max_length + 5  # 여유 공간 추가
                    ws.column_dimensions[column].width = adjusted_width
                    
                # K열에 1000단위 쉼표 추가
                for cell in ws['K']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = f"{int(cell.value):,}"
                        except ValueError:
                            try:
                                cell.value = f"{float(cell.value):,.2f}"
                            except ValueError:
                                pass
            
                # L열을 퍼센트 형식으로 표시
                for cell in ws['L']:
                    if cell.row == 1:
                        continue  # Skip header
                    if cell.value is not None:
                        try:
                            cell.value = float(cell.value) / 100  # 값을 100으로 나눔
                            cell.number_format = '0.00%'  # 퍼센트 형식 설정
                        except ValueError:
                            pass
            
                # N열에 하이퍼링크 추가
                for row in range(2, ws.max_row + 1):  # 데이터가 시작되는 2행부터 마지막 행까지
                    cell = ws[f'N{row}']  # N열의 셀
                    if cell.value:  # 셀에 값이 있는 경우
                        cell.hyperlink = cell.value  # 셀 값을 하이퍼링크로 설정
                        cell.style = "Hyperlink"  # 하이퍼링크 스타일 적용
            
                # 첫 행에 필터 추가
                ws.auto_filter.ref = ws.dimensions
                #---------------------------------------------------------------------------------------------------   
                wb.save(file_path)
                #QMessageBox.information(self, "성공", "데이터가 성공적으로 저장되었습니다!")
        except Exception as e:
                QMessageBox.critical(self, "오류", f"파일 저장 중 오류 발생: {str(e)}")

# PyQt5 애플리케이션 실행
QApplication.setStyle('Fusion')
app = QApplication(sys.argv)
main_dialog = MainDialog()
main_dialog.show()
sys.exit(app.exec_())